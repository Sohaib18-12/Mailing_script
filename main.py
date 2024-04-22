import pandas as pd
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from Body import create_message

workbook = load_workbook("Job_Track.xlsx")
data_sheet = workbook['Data']
tracking_sheet = workbook["Tracking"]

personal_info = pd.read_excel('Job_Track.xlsx', 'Personal Information')
df = pd.read_excel('Job_Track.xlsx', 'Data')
tracking = pd.read_excel('Job_Track.xlsx', 'Tracking')


def sendmail(to, position):

    fromaddr = personal_info.loc[0, "Email"]
    password = personal_info.loc[0, "Password"]
    toaddr = to

    # instance of MIMEMultipart
    msg = MIMEMultipart()

    # storing the senders email address
    msg['From'] = fromaddr

    # storing the receivers email address
    msg['To'] = toaddr

    # storing the subject
    msg['Subject'] = f"Application for {position} Position"

    # string to store the body of the mail
    with open("Body.html", "r", encoding='utf-8') as file:
        body = file.read()

    # attach the body with the msg instance
    msg.attach(MIMEText(body, 'html'))

    # open the file to be sent
    filename = personal_info.loc[0, "Resume"]
    attachment = open(f"./{personal_info.loc[0, 'Resume']}", "rb")

    # instance of MIMEBase and named as p
    p = MIMEBase('application', 'octet-stream')

    # To change the payload into encoded form
    p.set_payload(attachment.read())

    # encode into base64
    encoders.encode_base64(p)

    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    # attach the instance 'p' to instance 'msg'
    msg.attach(p)

    # Attach the cover letter
    filename = personal_info.loc[0, "Cover Letter"]
    attachment = open(f"./{personal_info.loc[0, 'Cover Letter']}", "rb")

    # instance of MIMEBase and named as p
    p = MIMEBase('application', 'octet-stream')

    # To change the payload into encoded form
    p.set_payload(attachment.read())

    # encode into base64
    encoders.encode_base64(p)

    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    # attach the instance 'p' to instance 'msg'
    msg.attach(p)

    # new
    # creates SMTP session
    s = smtplib.SMTP('smtp.gmail.com', 587)

    # start TLS for security
    s.starttls()

    # Authentication
    s.login(fromaddr, password)

    # Converts the Multipart msg into a string
    text = msg.as_string()

    # sending the mail
    s.sendmail(fromaddr, toaddr, text)

    # terminating the session
    s.quit()
    print("Sent...\n")


fullname = personal_info.loc[0, "First Name"] + " " + personal_info.loc[0, "Last Name"]

for (index, data) in df.iterrows():
    name = ""
    prefix = ""
    if data.Status != "Sent":
        if not pd.isnull(data.Email):
            position = data.Position
            company = data.Company

            if pd.isnull(data["First Name"]):
                name = "Hiring Team"
                prefix = ""
            else:
                name = data["First Name"]

                if data.Gender == "Male":
                    prefix = "Mr"
                elif data.Gender == "Female":
                    prefix = "Ms"
                else:
                    prefix = ""

            create_message(fullname, name, prefix, company, position)
            print(data.Company)

            sendmail(data.Email, position)

            data_sheet.cell(row=index+2, column=7, value="Sent")
            data.Status = "Sent"
            tracking_sheet.append(data.to_list())

        else:
            print(data.Company, ": Not Sent")
            data_sheet.cell(row=index + 2, column=7, value="Not Sent")

workbook.save("Job_Track.xlsx")

print("Finished")
